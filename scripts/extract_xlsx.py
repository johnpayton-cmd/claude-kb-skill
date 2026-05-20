"""Extract data from an XLSX file for KB summarization.

Usage:
  uv run --python 3.12 --with openpyxl extract_xlsx.py <path>
  uv run --python 3.12 --with openpyxl extract_xlsx.py <path> --sheet Sheet1
  uv run --python 3.12 --with openpyxl extract_xlsx.py <path> --headers-only
  uv run --python 3.12 --with openpyxl extract_xlsx.py <path> --max-rows 20

Workflow:
  1. Run without flags to see all sheet names and sample rows from each.
  2. Use --sheet to focus on a specific sheet.
  3. Use --headers-only to inspect column structure before extracting full content.
  4. Use --max-rows to control how much data is shown per sheet.
"""

import sys
import argparse
import openpyxl


def extract_xlsx(path, sheet_name=None, headers_only=False, max_rows=50):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sheet_names = wb.sheetnames
    print(f"Sheets ({len(sheet_names)}): {', '.join(sheet_names)}\n")

    target_sheets = [sheet_name] if sheet_name else sheet_names

    for name in target_sheets:
        if name not in sheet_names:
            print(f"Sheet '{name}' not found. Available: {', '.join(sheet_names)}", file=sys.stderr)
            continue

        ws = wb[name]
        print(f"=== {name} ===")

        headers = None
        row_count = 0
        for row in ws.iter_rows(values_only=True):
            cells = [str(c).strip() if c is not None else "" for c in row]
            if not any(cells):
                continue
            if headers is None:
                headers = cells
                print("Headers: " + " | ".join(headers))
                if headers_only:
                    break
                continue
            row_count += 1
            if row_count > max_rows:
                print(f"  ... (more rows — use --max-rows to see more)")
                break
            print(" | ".join(c[:200] for c in cells))

        print()

    wb.close()


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Extract data from an XLSX file for KB summarization."
    )
    parser.add_argument("path", help="Path to the XLSX file")
    parser.add_argument("--sheet", help="Extract only this sheet (default: all sheets)")
    parser.add_argument(
        "--headers-only",
        action="store_true",
        help="Show only column headers per sheet",
    )
    parser.add_argument(
        "--max-rows",
        type=int,
        default=50,
        help="Maximum rows to show per sheet (default: 50)",
    )
    args = parser.parse_args()

    extract_xlsx(
        args.path,
        sheet_name=args.sheet,
        headers_only=args.headers_only,
        max_rows=args.max_rows,
    )
